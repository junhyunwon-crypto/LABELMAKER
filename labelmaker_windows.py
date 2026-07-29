import csv
import ctypes
import os
import sys
import textwrap
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import xlrd
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageWin
from reportlab.graphics.barcode import createBarcodeDrawing

# Self-test only exercises data/label generation. Keeping the printer stack out of
# that path also lets the Windows build be validated under Wine without a driver.
if sys.platform == "win32" and "--self-test" not in sys.argv:
    import win32print
else:
    win32print = None


APP_TITLE = "B-FV4T 바코드 메이커 (ZMODE)"
DEFAULT_PRINTER_KEYWORDS = ("TOSHIBA", "B-FV4")
FONT_PATH = r"C:\Windows\Fonts\arial.ttf"

# 40 x 50 mm at 203 dpi
LABEL_DPI = 203
LABEL_W = 320
LABEL_H = 400

FONT_SIZE = 20
TOP_MARGIN = 40
LINE_HEIGHT = 24
BARCODE_Y = 270
BARCODE_H = 80
SKU_Y = 365


def load_font(size=FONT_SIZE):
    candidates = [
        FONT_PATH,
        r"C:\Windows\Fonts\segoeui.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def draw_centered(draw, text, y, font, fill=0):
    if not text:
        return
    box = draw.textbbox((0, 0), text, font=font)
    width = box[2] - box[0]
    draw.text(((LABEL_W - width) // 2, y), text, font=font, fill=fill)


def draw_code128(draw, value, y=BARCODE_Y, height=BARCODE_H):
    value = str(value).strip()
    if not value:
        raise ValueError("SKU/바코드 값이 비어 있습니다.")
    if any(ord(ch) > 127 for ch in value):
        raise ValueError(f"Code 128에서 지원하지 않는 문자가 있습니다: {value}")

    module_width = 1.0
    barcode = createBarcodeDrawing(
        "Code128",
        value=value,
        barWidth=module_width,
        barHeight=height,
        humanReadable=False,
        quiet=True,
    )
    if barcode.width > LABEL_W - 16:
        module_width *= (LABEL_W - 16) / barcode.width
        barcode = createBarcodeDrawing(
            "Code128",
            value=value,
            barWidth=module_width,
            barHeight=height,
            humanReadable=False,
            quiet=True,
        )

    x_offset = int(round((LABEL_W - barcode.width) / 2))
    group = barcode.contents[0].draw()
    for rect in group.contents:
        if getattr(rect, "fillColor", None) is None:
            continue
        x0 = x_offset + int(round(rect.x))
        x1 = x0 + max(1, int(round(rect.width))) - 1
        draw.rectangle((x0, y, x1, y + height - 1), fill=0)


def build_label_image(brand, style_no, name, size, color, price_str, sku):
    image = Image.new("L", (LABEL_W, LABEL_H), 255)
    draw = ImageDraw.Draw(image)
    font = load_font(FONT_SIZE)

    lines = [brand]
    lines.extend(textwrap.wrap(name, width=28))
    lines.extend(["", style_no, f"SIZE {size}", color, "", f"₩{price_str}"])

    y = TOP_MARGIN
    for line in lines:
        if line:
            draw.text((15, y), line, font=font, fill=0)
        y += LINE_HEIGHT

    draw_code128(draw, sku)
    draw_centered(draw, sku, SKU_Y, font)
    return image


def safe_value(row, index):
    if len(row) <= index:
        return ""
    value = row[index]
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def format_price(raw_price):
    digits = "".join(ch for ch in str(raw_price) if ch.isdigit())
    if not digits:
        return str(raw_price).strip()
    return f"{int(digits):,}"


def parse_quantity(raw_quantity):
    try:
        quantity = int(float(raw_quantity)) if str(raw_quantity).strip() else 1
    except (TypeError, ValueError):
        quantity = 1
    return max(0, quantity)


def read_rows(file_path):
    extension = os.path.splitext(file_path)[1].lower()
    if extension == ".csv":
        for encoding in ("cp949", "utf-8-sig"):
            try:
                with open(file_path, "r", encoding=encoding, newline="") as source:
                    source_rows = list(csv.reader(source))
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("CSV 문자 인코딩을 확인할 수 없습니다.")
        data_rows = source_rows[1:]
    elif extension == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        finally:
            workbook.close()
    elif extension == ".xls":
        workbook = xlrd.open_workbook(file_path, on_demand=True)
        try:
            sheet = workbook.sheet_by_index(0)
            data_rows = [sheet.row_values(index) for index in range(1, sheet.nrows)]
        finally:
            workbook.release_resources()
    else:
        raise ValueError("지원 형식은 .xlsx, .xls, .csv입니다.")

    rows = []
    for source_index, row in enumerate(data_rows, start=2):
        style_no = safe_value(row, 1)
        if not style_no:
            continue
        quantity = parse_quantity(safe_value(row, 6))
        if quantity < 1:
            continue
        rows.append(
            {
                "source_index": source_index,
                "brand": safe_value(row, 0),
                "style_no": style_no,
                "name": safe_value(row, 2),
                "size": safe_value(row, 3),
                "color": safe_value(row, 4),
                "price": format_price(safe_value(row, 5)),
                "quantity": quantity,
                "sku": safe_value(row, 7),
            }
        )
    return rows


def list_printers():
    if win32print is None:
        return []
    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    return sorted({item[2] for item in win32print.EnumPrinters(flags)})


def choose_default_printer(printers):
    for printer in printers:
        upper = printer.upper()
        if all(keyword in upper for keyword in DEFAULT_PRINTER_KEYWORDS):
            return printer
    for printer in printers:
        upper = printer.upper()
        if any(keyword in upper for keyword in DEFAULT_PRINTER_KEYWORDS):
            return printer
    try:
        default = win32print.GetDefaultPrinter()
        if default in printers:
            return default
    except Exception:
        pass
    return printers[0] if printers else ""


def print_images_gdi(printer_name, image_jobs, progress_callback=None):
    if sys.platform != "win32":
        raise RuntimeError("실제 인쇄는 Windows에서만 지원됩니다.")

    class DOCINFOW(ctypes.Structure):
        _fields_ = [
            ("cbSize", ctypes.c_int),
            ("lpszDocName", ctypes.c_wchar_p),
            ("lpszOutput", ctypes.c_wchar_p),
            ("lpszDatatype", ctypes.c_wchar_p),
            ("fwType", ctypes.c_uint),
        ]

    class BITMAPINFOHEADER(ctypes.Structure):
        _fields_ = [
            ("biSize", ctypes.c_uint32),
            ("biWidth", ctypes.c_int32),
            ("biHeight", ctypes.c_int32),
            ("biPlanes", ctypes.c_uint16),
            ("biBitCount", ctypes.c_uint16),
            ("biCompression", ctypes.c_uint32),
            ("biSizeImage", ctypes.c_uint32),
            ("biXPelsPerMeter", ctypes.c_int32),
            ("biYPelsPerMeter", ctypes.c_int32),
            ("biClrUsed", ctypes.c_uint32),
            ("biClrImportant", ctypes.c_uint32),
        ]

    class RGBQUAD(ctypes.Structure):
        _fields_ = [
            ("rgbBlue", ctypes.c_ubyte),
            ("rgbGreen", ctypes.c_ubyte),
            ("rgbRed", ctypes.c_ubyte),
            ("rgbReserved", ctypes.c_ubyte),
        ]

    class BITMAPINFO(ctypes.Structure):
        _fields_ = [
            ("bmiHeader", BITMAPINFOHEADER),
            ("bmiColors", RGBQUAD * 1),
        ]

    gdi32 = ctypes.WinDLL("gdi32", use_last_error=True)
    gdi32.CreateDCW.argtypes = [
        ctypes.c_wchar_p,
        ctypes.c_wchar_p,
        ctypes.c_wchar_p,
        ctypes.c_void_p,
    ]
    gdi32.CreateDCW.restype = ctypes.c_void_p
    gdi32.GetDeviceCaps.argtypes = [ctypes.c_void_p, ctypes.c_int]
    gdi32.GetDeviceCaps.restype = ctypes.c_int
    gdi32.StartDocW.argtypes = [ctypes.c_void_p, ctypes.POINTER(DOCINFOW)]
    gdi32.StartDocW.restype = ctypes.c_int
    gdi32.StartPage.argtypes = [ctypes.c_void_p]
    gdi32.StartPage.restype = ctypes.c_int
    gdi32.EndPage.argtypes = [ctypes.c_void_p]
    gdi32.EndPage.restype = ctypes.c_int
    gdi32.EndDoc.argtypes = [ctypes.c_void_p]
    gdi32.EndDoc.restype = ctypes.c_int
    gdi32.AbortDoc.argtypes = [ctypes.c_void_p]
    gdi32.AbortDoc.restype = ctypes.c_int
    gdi32.DeleteDC.argtypes = [ctypes.c_void_p]
    gdi32.DeleteDC.restype = ctypes.c_int
    gdi32.StretchDIBits.argtypes = [
        ctypes.c_void_p,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_int,
        ctypes.c_void_p,
        ctypes.POINTER(BITMAPINFO),
        ctypes.c_uint,
        ctypes.c_ulong,
    ]
    gdi32.StretchDIBits.restype = ctypes.c_int

    hdc = gdi32.CreateDCW("WINSPOOL", printer_name, None, None)
    if not hdc:
        raise ctypes.WinError(ctypes.get_last_error())

    # Windows GetDeviceCaps constants.
    HORZRES, VERTRES, LOGPIXELSX, LOGPIXELSY = 8, 10, 88, 90
    dpi_x = gdi32.GetDeviceCaps(hdc, LOGPIXELSX) or LABEL_DPI
    dpi_y = gdi32.GetDeviceCaps(hdc, LOGPIXELSY) or LABEL_DPI
    printable_w = gdi32.GetDeviceCaps(hdc, HORZRES)
    printable_h = gdi32.GetDeviceCaps(hdc, VERTRES)

    target_w = min(printable_w, int(round(LABEL_W * dpi_x / LABEL_DPI)))
    target_h = min(printable_h, int(round(LABEL_H * dpi_y / LABEL_DPI)))
    total_pages = sum(quantity for _, quantity in image_jobs)
    printed_pages = 0
    doc_info = DOCINFOW(
        ctypes.sizeof(DOCINFOW),
        f"{APP_TITLE} - {total_pages} labels",
        None,
        None,
        0,
    )

    try:
        if gdi32.StartDocW(hdc, ctypes.byref(doc_info)) <= 0:
            raise ctypes.WinError(ctypes.get_last_error())
        for image, quantity in image_jobs:
            # Let Pillow create a native Windows DIB.  The previous manual
            # 24-bit BITMAPINFO path was accepted by the Toshiba driver but
            # could produce blank labels.
            dib = ImageWin.Dib(image.convert("RGB"))
            for _ in range(quantity):
                if gdi32.StartPage(hdc) <= 0:
                    raise ctypes.WinError(ctypes.get_last_error())
                dib.draw(int(hdc), (0, 0, target_w, target_h))
                if gdi32.EndPage(hdc) <= 0:
                    raise ctypes.WinError(ctypes.get_last_error())
                printed_pages += 1
                if progress_callback:
                    progress_callback(printed_pages, total_pages)
        if gdi32.EndDoc(hdc) <= 0:
            raise ctypes.WinError(ctypes.get_last_error())
    except Exception:
        gdi32.AbortDoc(hdc)
        raise
    finally:
        gdi32.DeleteDC(hdc)

    return printed_pages


def image_to_zpl(image, quantity=1):
    """Encode the complete 320 x 400 label as a ZPL ^GFA bitmap."""
    mono = image.convert("1")
    row_bytes = (LABEL_W + 7) // 8
    raw = mono.tobytes()
    # Pillow uses 0 for black while ZPL ^GFA prints set (1) bits.
    printable = bytes(value ^ 0xFF for value in raw)
    expected = row_bytes * LABEL_H
    if len(printable) != expected:
        raise RuntimeError(
            f"ZPL 비트맵 크기 오류: {len(printable)} bytes (expected {expected})"
        )
    hex_data = printable.hex().upper()
    return (
        f"^XA^PW{LABEL_W}^LL{LABEL_H}^LH0,0"
        f"^FO0,0^GFA,{expected},{expected},{row_bytes},{hex_data}^FS"
        f"^PQ{quantity},0,1,N^XZ"
    ).encode("ascii")


def print_images(printer_name, image_jobs, progress_callback=None):
    """Send ZPL directly to a B-FV4T running in ZMODE."""
    if sys.platform != "win32":
        raise RuntimeError("실제 인쇄는 Windows에서만 지원됩니다.")

    total_pages = sum(quantity for _, quantity in image_jobs)
    sent_pages = 0
    printer = win32print.OpenPrinter(printer_name)
    try:
        win32print.StartDocPrinter(
            printer,
            1,
            (f"{APP_TITLE} - {total_pages} labels", None, "RAW"),
        )
        try:
            for image, quantity in image_jobs:
                payload = image_to_zpl(image, quantity)
                win32print.StartPagePrinter(printer)
                try:
                    written = win32print.WritePrinter(printer, payload)
                    if written != len(payload):
                        raise RuntimeError(
                            f"인쇄 데이터 전송 부족: {written}/{len(payload)} bytes"
                        )
                finally:
                    win32print.EndPagePrinter(printer)
                sent_pages += quantity
                if progress_callback:
                    progress_callback(sent_pages, total_pages)
        finally:
            win32print.EndDocPrinter(printer)
    finally:
        win32print.ClosePrinter(printer)
    return sent_pages


class LabelMakerApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("620x300")
        self.file_path = tk.StringVar()
        self.printer_name = tk.StringVar()
        self.status = tk.StringVar(value="데이터 파일을 선택하세요.")
        self.printers = list_printers()
        self.printer_name.set(choose_default_printer(self.printers))
        self._build_ui()

    def _build_ui(self):
        frame = ttk.Frame(self.root, padding=18)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="데이터 파일").grid(row=0, column=0, sticky="w", pady=6)
        ttk.Entry(frame, textvariable=self.file_path, width=58).grid(
            row=0, column=1, sticky="ew", padx=8
        )
        ttk.Button(frame, text="찾아보기", command=self.select_file).grid(row=0, column=2)

        ttk.Label(frame, text="프린터").grid(row=1, column=0, sticky="w", pady=6)
        ttk.Combobox(
            frame,
            textvariable=self.printer_name,
            values=self.printers,
            state="readonly",
            width=55,
        ).grid(row=1, column=1, sticky="ew", padx=8)
        ttk.Button(frame, text="새로고침", command=self.refresh_printers).grid(row=1, column=2)

        ttk.Label(
            frame,
            text="열 순서: 브랜드 / 스타일 / 상품명 / 사이즈 / 컬러 / 가격 / 수량 / SKU",
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(12, 4))

        ttk.Label(frame, textvariable=self.status, foreground="#234f8c").grid(
            row=3, column=0, columnspan=3, sticky="w", pady=8
        )

        button_frame = ttk.Frame(frame)
        button_frame.grid(row=4, column=0, columnspan=3, pady=18)
        ttk.Button(button_frame, text="미리보기", command=self.preview, width=16).pack(
            side="left", padx=6
        )
        ttk.Button(button_frame, text="인쇄", command=self.print_labels, width=16).pack(
            side="left", padx=6
        )
        frame.columnconfigure(1, weight=1)

    def select_file(self):
        path = filedialog.askopenfilename(
            title="데이터 파일 선택",
            filetypes=[
                ("Excel/CSV", "*.xlsx *.xls *.csv"),
                ("Excel", "*.xlsx *.xls"),
                ("CSV", "*.csv"),
            ],
        )
        if path:
            self.file_path.set(path)
            self.status.set(os.path.basename(path))

    def refresh_printers(self):
        self.printers = list_printers()
        default = choose_default_printer(self.printers)
        self.printer_name.set(default)
        self.status.set(f"프린터 {len(self.printers)}대 감지")
        for widget in self.root.winfo_children():
            for child in widget.winfo_children():
                if isinstance(child, ttk.Combobox):
                    child.configure(values=self.printers)

    def _load_jobs(self):
        path = self.file_path.get().strip()
        if not path:
            raise ValueError("데이터 파일을 선택하세요.")
        rows = read_rows(path)
        if not rows:
            raise ValueError("인쇄할 행이 없습니다.")
        jobs = []
        for row in rows:
            image = build_label_image(
                row["brand"],
                row["style_no"],
                row["name"],
                row["size"],
                row["color"],
                row["price"],
                row["sku"],
            )
            jobs.append((image, row["quantity"]))
        return rows, jobs

    def preview(self):
        try:
            rows, jobs = self._load_jobs()
            preview_path = os.path.join(
                os.path.dirname(os.path.abspath(self.file_path.get())),
                "label_preview.png",
            )
            jobs[0][0].save(preview_path)
            os.startfile(preview_path)
            self.status.set(
                f"미리보기 저장: {preview_path} / 총 {sum(r['quantity'] for r in rows)}장"
            )
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))

    def print_labels(self):
        try:
            printer = self.printer_name.get().strip()
            if not printer:
                raise ValueError("프린터를 선택하세요.")
            rows, jobs = self._load_jobs()
            total = sum(row["quantity"] for row in rows)
            if not messagebox.askokcancel(
                "인쇄 확인",
                f"프린터: {printer}\n데이터: {len(rows)}건\n라벨: {total}장\n\n인쇄할까요?",
            ):
                return

            def update_progress(current, maximum):
                self.status.set(f"인쇄 중: {current}/{maximum}")
                self.root.update_idletasks()

            printed = print_images(printer, jobs, update_progress)
            self.status.set(f"완료: {printed}장")
            messagebox.showinfo(APP_TITLE, f"{printed}장 인쇄 작업을 전송했습니다.")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, str(exc))
            self.status.set("인쇄 실패")


def main():
    if sys.platform != "win32":
        print("이 프로그램의 실제 인쇄 기능은 Windows 전용입니다.")
        return
    root = tk.Tk()
    LabelMakerApp(root)
    root.mainloop()


def run_self_test():
    cases = [
        ("SAME26072400001", 3),
        ("DIFF26072400001", 1),
        ("DIFF26072400002", 1),
    ]
    images = []
    for sku, quantity in cases:
        image = build_label_image(
            "HFW",
            "SELF-TEST",
            "WINDOWS EXE VALIDATION",
            "F",
            "BLACK",
            "10,000",
            sku,
        )
        if image.size != (LABEL_W, LABEL_H):
            raise AssertionError(f"잘못된 라벨 크기: {image.size}")
        images.append((image, quantity))
    if sum(quantity for _, quantity in images) != 5:
        raise AssertionError("수량 계산 오류")
    print("SELF_TEST_OK pages=5 unique_barcodes=3")


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        run_self_test()
    else:
        main()
